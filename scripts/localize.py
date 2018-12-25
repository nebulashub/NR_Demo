#!/usr/bin/env python3
# - *- coding: utf- 8 - *-

import os
import sys
from openpyxl import load_workbook


zh_dir = "zh-Hans.lproj"
en_dir = "en.lproj"
kr_dir = "ko-KR.lproj"
string_name = "Localizable.strings"

os.chdir(sys.path[0])
root_dir = os.path.abspath("../")

output_dir = os.path.join(root_dir, "src", "pages", "index", "strings")

o_zh_path = os.path.join(output_dir, "cn.js")
o_en_path = os.path.join(output_dir, "en.js")


def write_resource(file_name, kv):
    s = "export default {\n"
    i = 0
    for k in sorted(kv.keys()):
        s += "  " + k + ": '" + kv[k].replace("\\", "\\\\").replace("\n", "\\n").replace("\'", "\\\'") + "'"
        if i < len(kv) - 1:
            s += ",\n"
        else:
            s += "\n"
        i += 1
    s += "}\n"
    f = open(file_name, "w", encoding="utf-8")
    f.write(s)
    f.close()

def process_excel(excel_file):
    wb = load_workbook(excel_file)
    zh_kv = {}
    en_kv = {}
    all_key = []
    ws = wb['NR demo']
    col_zh, col_en, col_key = get_columns(ws)
    for i in range(2, 500):
        keys = ws[col_key + str(i)].value
        if keys is not None:
            keys.replace(" ", "")
            ks = keys.strip().split(",")
            for key in ks:
                if key in all_key:
                    print("key 重复：" + key)
                else:
                    all_key.append(key)
                zh = ws[col_zh + str(i)].value
                en = ws[col_en + str(i)].value
                if zh is not None:
                    zh_kv[key] = zh
                if en is not None:
                    en_kv[key] = en
    return zh_kv, en_kv


def get_columns(ws):
    col_zh = ""
    col_en = ""
    col_key = ""
    for i in range(0, 26):
        c = chr(ord('A') + i)
        cname = ws[c + str(1)].value
        if cname == "中":
            col_zh = c
        elif cname == "英":
            col_en = c
        elif cname == "key":
            col_key = c
    return col_zh, col_en, col_key


def check_dir(dir_path):
    b = os.path.exists(dir_path)
    if not b:
        os.makedirs(dir_path)


def init_env():
    check_dir(output_dir)
    return True


def main(excel_path):
    if not os.path.exists(excel_path):
        print("找不到 excel 文件")
        return
    init_env()

    excel_kvs = process_excel(excel_path)
    write_resource(o_zh_path, excel_kvs[0])
    write_resource(o_en_path, excel_kvs[1])
    print("完成")


if __name__ == '__main__':
    if len(sys.argv) == 1:
        print("错误：请传入 excel 文件路径")
    else:
        main(sys.argv[1])
